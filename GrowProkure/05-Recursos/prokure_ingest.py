#!/usr/bin/env python3
"""
Prokure — motor de ingesta del _INBOX.
Lee cada Excel nuevo en 02-Investigacion/_INBOX/, normaliza, DEDUPLICA contra la
Base Maestra, integra solo lo nuevo, mueve el archivo a _INBOX/_procesados/ y reporta.

Uso: python3 prokure_ingest.py
Detecta columnas por nombre (ES/EN). Si un archivo no se puede mapear, lo reporta
y lo deja en el INBOX para revisión manual (el agente Prokure lo mapea a mano).
"""
import os, shutil, datetime
import openpyxl

BASE = "/home/user/Cosas-Carlos/GrowProkure"
INBOX = f"{BASE}/02-Investigacion/_INBOX"
DONE = f"{INBOX}/_procesados"
MASTER = f"{BASE}/05-Recursos/Base_Maestra_GrowProkure.xlsx"

COLS = ["Vertical","Pais","Lado","Empresa","Contacto","Puesto","Funcion","Ciudad",
        "Estado","Email","Telefono","Segmento","Prioridad","LinkedIn","Fuente","Origen"]

# sinónimos de encabezado -> campo canónico
SYN = {
 "email":"Email","correo":"Email","e-mail":"Email",
 "empresa":"Empresa","company":"Empresa","cuenta":"Empresa","empresa_actual":"Empresa",
 "contacto":"Contacto","contact":"Contacto","nombre":"Contacto","name":"Contacto",
 "puesto":"Puesto","title":"Puesto","cargo":"Puesto","puesto_actual":"Puesto",
 "funcion":"Funcion","función":"Funcion","target function":"Funcion","perfil":"Funcion","función objetivo":"Funcion","funcion objetivo":"Funcion",
 "ciudad":"Ciudad","city":"Ciudad","estado":"Estado","state":"Estado",
 "telefono":"Telefono","teléfono":"Telefono","phone":"Telefono","tel directo":"Telefono","direct phone":"Telefono","celular":"Telefono","mobile":"Telefono",
 "segmento":"Segmento","segment":"Segmento","industria":"Segmento",
 "prioridad":"Prioridad","priority":"Prioridad","linkedin":"LinkedIn",
 "vertical":"Vertical","pais":"Pais","país":"Pais","country":"Pais",
 "fuente":"Fuente","source":"Fuente","lado":"Lado","side":"Lado",
}

def norm_email(e):
    if not e: return ""
    e=str(e).strip().lower(); return e if "@" in e else ""

def find_header(ws, maxscan=6):
    """Devuelve (fila_header, {col_idx: campo}) buscando la fila con más columnas conocidas."""
    best=(None,{},0)
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=maxscan,values_only=True),1):
        mapping={}
        for idx,val in enumerate(row):
            if val is None: continue
            key=str(val).strip().lower()
            if key in SYN: mapping[idx]=SYN[key]
        if len(mapping)>best[2]: best=(i,mapping,len(mapping))
    return best[0],best[1]

def load_master():
    wb=openpyxl.load_workbook(MASTER)
    ws=wb["Contactos"]
    emails=set(); namecomp=set()
    for r in ws.iter_rows(min_row=2,values_only=True):
        d=dict(zip(COLS,r))
        if d.get("Email"): emails.add(str(d["Email"]).strip().lower())
        if d.get("Contacto"): namecomp.add((str(d["Contacto"]).strip().lower(),str(d.get("Empresa","")).strip().lower()))
    return wb,ws,emails,namecomp

def process_file(path, ws, emails, namecomp, origin):
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
    added=0; dup=0; scanned=0; unmapped_sheets=0
    for sh in wb.worksheets:
        hrow,mapping=find_header(sh)
        if not mapping or "Email" not in mapping.values() and "Contacto" not in mapping.values():
            unmapped_sheets+=1; continue
        for r in sh.iter_rows(min_row=hrow+1,values_only=True):
            rec={c:"" for c in COLS}
            for idx,field in mapping.items():
                if idx<len(r) and r[idx] is not None:
                    rec[field]=str(r[idx]).strip()
            if not rec["Contacto"] and not rec["Email"]: continue
            scanned+=1
            rec["Email"]=norm_email(rec["Email"])
            key_e=rec["Email"]; key_n=(rec["Contacto"].lower(),rec["Empresa"].lower())
            if key_e and key_e in emails: dup+=1; continue
            if not key_e and rec["Contacto"] and key_n in namecomp: dup+=1; continue
            if key_e: emails.add(key_e)
            if rec["Contacto"]: namecomp.add(key_n)
            if not rec["Origen"]: rec["Origen"]=origin
            if not rec["Fuente"]: rec["Fuente"]=origin
            ws.append([rec[c] for c in COLS]); added+=1
    wb.close()
    return dict(scanned=scanned, added=added, dup=dup, unmapped=unmapped_sheets)

def main():
    os.makedirs(DONE, exist_ok=True)
    files=[f for f in os.listdir(INBOX)
           if f.lower().endswith((".xlsx",".xls")) and not f.startswith("~")]
    if not files:
        print("INBOX vacío — nada que integrar."); return
    wb,ws,emails,namecomp=load_master()
    print(f"Base maestra actual: {ws.max_row-1} contactos.\n")
    total_new=0
    for f in files:
        p=os.path.join(INBOX,f)
        rep=process_file(p, ws, emails, namecomp, origin=f"INBOX:{f}")
        total_new+=rep["added"]
        print(f"### {f}")
        print(f"  escaneados: {rep['scanned']} | NUEVOS: {rep['added']} | duplicados: {rep['dup']} | hojas sin mapear: {rep['unmapped']}")
        if rep["added"]>0 or rep["scanned"]>0:
            shutil.move(p, os.path.join(DONE, f))
            print(f"  -> movido a _procesados/")
        else:
            print(f"  -> se deja en INBOX (revisar mapeo manual)")
    wb.save(MASTER)
    print(f"\n✅ Integrados {total_new} contactos nuevos. Base maestra: {ws.max_row-1} contactos.")

if __name__=="__main__":
    main()
