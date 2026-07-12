#!/usr/bin/env python3
"""Consolida los 3 estudios en una base maestra GrowProkure (Capa 1)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = "/home/user/Cosas-Carlos/GrowProkure/02-Investigacion"
MX = f"{SRC}/electronica/Estudio_Electronica_Mexico.xlsx"
US = f"{SRC}/electronica/US_Electronics_SMT_Study.xlsx"
FOIL = f"{SRC}/plasticos/Contactos_Foil_TID_Global.xlsx"
OUT = f"{SRC}/../05-Recursos/Base_Maestra_GrowProkure.xlsx"

def norm_email(e):
    if not e: return ""
    e = str(e).strip().lower()
    return e if "@" in e else ""

contacts = []  # unified dicts

def add(vertical, pais, lado, empresa, contacto, puesto, funcion, ciudad, estado,
        email, tel, segmento, prioridad, linkedin, fuente, origen):
    contacts.append(dict(
        Vertical=vertical, Pais=pais, Lado=lado, Empresa=(empresa or "").strip(),
        Contacto=(contacto or "").strip(), Puesto=(puesto or "").strip(),
        Funcion=(funcion or "").strip(), Ciudad=(ciudad or "").strip(),
        Estado=(estado or "").strip(), Email=norm_email(email),
        Telefono=(str(tel).strip() if tel else ""), Segmento=(segmento or "").strip(),
        Prioridad=(prioridad or "").strip(), LinkedIn=(linkedin or "").strip(),
        Fuente=(fuente or "").strip(), Origen=origen))

# ---- MX Electronica: Contactos ----
wb = openpyxl.load_workbook(MX, read_only=True, data_only=True)
for r in wb["Contactos"].iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    add("Electronica","MX","Demanda", r[0], r[5], r[6], r[4], r[2], r[3],
        r[8], r[9] or r[10], "", "", "", r[16], "Estudio MX")
wb.close()

# ---- US SMT: Contacts ----
wb = openpyxl.load_workbook(US, read_only=True, data_only=True)
for r in wb["Contacts"].iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    add("Electronica","US","Demanda", r[0], r[5], r[6], r[4], r[2], r[3],
        r[8], r[9] or r[10], "", "", "", r[16], "Estudio US-SMT")
wb.close()

# ---- FOIL: Prosp Todos (header row 3) ----
wb = openpyxl.load_workbook(FOIL, read_only=True, data_only=True)
for r in wb["Prosp — Todos"].iter_rows(min_row=4, values_only=True):
    if not r[1]: continue  # Empresa
    add("Plasticos","MX","Demanda", r[1], r[2], r[3], r[4], r[5], r[6],
        r[9], r[10], r[0], r[12], r[8], r[11], "Foil TID - Prospeccion")
# Agenda propia (header row 3): 0 Contacto,1 Puesto,2 Emp_estudio,3 Emp_actual,4 Puesto_actual,5 Estatus,6 Email,7 Tel,8 Fuentes
for r in wb["Agenda propia"].iter_rows(min_row=4, values_only=True):
    if not r[0]: continue
    add("Plasticos","MX","Demanda", r[3] or r[2], r[0], r[4] or r[1], "Agenda", "", "",
        r[6], r[7], "", "Caliente" if str(r[5]).upper()=="VIGENTE" else "", "", r[8], "Foil TID - Agenda")
wb.close()

# ---- DEDUP ----
seen_email, seen_namecomp = {}, {}
unique, dups = [], 0
for c in contacts:
    key_e = c["Email"]
    key_n = (c["Contacto"].lower(), c["Empresa"].lower())
    if key_e and key_e in seen_email:
        dups += 1; continue
    if not key_e and c["Contacto"] and key_n in seen_namecomp:
        dups += 1; continue
    if key_e: seen_email[key_e] = True
    if c["Contacto"]: seen_namecomp[key_n] = True
    unique.append(c)

# ---- EMPRESAS master from directories ----
empresas = {}
def add_emp(vertical, pais, empresa, tipo, ciudad, estado, sector, evidencia, lineas, conf, web, fuente):
    if not empresa: return
    k = (empresa.strip().lower(), pais)
    if k in empresas: return
    empresas[k] = dict(Vertical=vertical, Pais=pais, Empresa=empresa.strip(), Tipo=(tipo or ""),
        Ciudad=(ciudad or ""), Estado=(estado or ""), Sector=(sector or ""),
        Evidencia=(evidencia or ""), LineasSMT=(lineas or ""), Confianza=(conf or ""),
        Web=(web or ""), Fuente=(fuente or ""), Lado="Demanda")

wb = openpyxl.load_workbook(MX, read_only=True, data_only=True)
for r in wb["Directorio"].iter_rows(min_row=2, values_only=True):
    add_emp("Electronica","MX", r[2], r[3], r[4], r[5], r[7], r[8], r[9], r[10], r[11], r[12])
wb.close()
wb = openpyxl.load_workbook(US, read_only=True, data_only=True)
for r in wb["Directory"].iter_rows(min_row=2, values_only=True):
    add_emp("Electronica","US", r[2], r[3], r[4], r[5], r[7], r[8], r[9], r[10], r[11], r[12])
wb.close()
# foil empresas from prospection
wb = openpyxl.load_workbook(FOIL, read_only=True, data_only=True)
for r in wb["Prosp — Todos"].iter_rows(min_row=4, values_only=True):
    if r[1]: add_emp("Plasticos","MX", r[1], "", r[5], r[6], r[7], "", "", "", "", "Foil TID")
wb.close()

# ---- WRITE ----
out = openpyxl.Workbook()
HEAD = PatternFill("solid", fgColor="B45F27"); HF = Font(color="FFFFFF", bold=True)
def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c); cell.fill=HEAD; cell.font=HF
        cell.alignment=Alignment(vertical="center")
    ws.freeze_panes="A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

# Contactos sheet
ws = out.active; ws.title="Contactos"
cols = ["Vertical","Pais","Lado","Empresa","Contacto","Puesto","Funcion","Ciudad","Estado","Email","Telefono","Segmento","Prioridad","LinkedIn","Fuente","Origen"]
ws.append(cols)
for c in unique: ws.append([c[k] for k in cols])
style_header(ws, len(cols))
for i,w in enumerate([12,6,10,26,22,26,14,18,14,30,16,12,10,28,16,20],1):
    ws.column_dimensions[get_column_letter(i)].width=w

# Empresas sheet
ws2 = out.create_sheet("Empresas")
cols2=["Vertical","Pais","Lado","Empresa","Tipo","Ciudad","Estado","Sector","Evidencia","LineasSMT","Confianza","Web","Fuente"]
ws2.append(cols2)
for e in empresas.values(): ws2.append([e[k] for k in cols2])
style_header(ws2, len(cols2))
for i,w in enumerate([12,6,10,30,16,18,14,22,26,10,10,26,16],1):
    ws2.column_dimensions[get_column_letter(i)].width=w

# Resumen
ws3 = out.create_sheet("Resumen", 0)
def emails(v=None,p=None):
    return sum(1 for c in unique if c["Email"] and (v is None or c["Vertical"]==v) and (p is None or c["Pais"]==p))
def cnt(v=None,p=None):
    return sum(1 for c in unique if (v is None or c["Vertical"]==v) and (p is None or c["Pais"]==p))
rows=[
 ["BASE MAESTRA GrowProkure — Capa 1 (Datos)",""],
 ["Generado","2026-07-12"],
 ["",""],
 ["CONTACTOS (deduplicados)",""],
 ["  Total contactos únicos", len(unique)],
 ["  Con email", emails()],
 ["  Duplicados removidos", dups],
 ["",""],
 ["  Electrónica MX", f"{cnt('Electronica','MX')} ({emails('Electronica','MX')} email)"],
 ["  Electrónica US", f"{cnt('Electronica','US')} ({emails('Electronica','US')} email)"],
 ["  Plásticos/Foil MX", f"{cnt('Plasticos','MX')} ({emails('Plasticos','MX')} email)"],
 ["",""],
 ["EMPRESAS (únicas)", len(empresas)],
 ["  Electrónica", sum(1 for e in empresas.values() if e['Vertical']=='Electronica')],
 ["  Plásticos", sum(1 for e in empresas.values() if e['Vertical']=='Plasticos')],
 ["",""],
 ["Hojas","Contactos, Empresas, Resumen"],
 ["Nota","Lado=Demanda (compradores objetivo). Oferta=clientes GrowProkure (Astute/TID)."],
]
for r in rows: ws3.append(r)
ws3.cell(1,1).font=Font(bold=True,size=14,color="B45F27")
ws3.column_dimensions["A"].width=34; ws3.column_dimensions["B"].width=44

out.save(OUT)
print(f"OK -> {OUT}")
print(f"Contactos únicos: {len(unique)} | con email: {emails()} | dups removidos: {dups}")
print(f"Empresas únicas: {len(empresas)}")
