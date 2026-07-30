#!/usr/bin/env python3
"""
whoop_dashboard.py — Genera un dashboard estetico en Excel con los datos de WHOOP.

Lee el JSON crudo que dejo whoop_sync.py en ./datos/ y arma un archivo Excel con:
  - Portada "Dashboard": tarjetas KPI (recovery, HRV, FC reposo, sueno, strain) con
    colores por zona, y graficas de tendencia (recovery y HRV).
  - Hojas de detalle: Recovery, Sueno, Strain (cycles) y Workouts, con formato.

Uso:
    python whoop_sync.py        # (1) baja/actualiza los datos
    python whoop_dashboard.py   # (2) genera el Excel
Luego abrelo con:  open TID-MAX-WHOOP.xlsx   (en Mac)

Opcional:
    python whoop_dashboard.py [carpeta_datos] [archivo_salida.xlsx]
"""
import os
import sys
import glob
import json
import csv
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Falta 'openpyxl'. Instala con:  pip install -r requirements.txt")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "datos")
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, "TID-MAX-WHOOP.xlsx")

# --- Paleta (oscura con acento teal, estilo TID-MAX) ---
INK = "0B1220"       # header casi negro
PANEL = "16233A"     # paneles
ACCENT = "16E0B5"    # teal
BLUE = "3DA5FF"
INDIGO = "7C6CF6"
ORANGE = "FF8A3D"
GOOD = "1DB954"      # verde
WARN = "F5B301"      # ambar
BAD = "E5484D"       # rojo
WHITE = "FFFFFF"
MUTED = "9FB3C8"
ROW_ALT = "F3F6FA"
GRID = "D6DEE8"
CYAN = "0EA5E9"      # natacion (agua)
AQUA = "0284C7"
TEAL = "06B6D4"
SLATE = "64748B"
# Semaforo pastel (tonos tenues)
GREEN_SOFT = "D6EFD5"
YELLOW_SOFT = "FCF2C9"
RED_SOFT = "F9D6D2"


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _light_gridlines():
    gl = ChartLines()
    gl.spPr = GraphicalProperties()
    gl.spPr.line = LineProperties(solidFill="E1E8F0", w=6350)  # gris claro, delgado
    return gl


def style_chart(chart, n, series_color):
    """Ejes visibles con numeros, rejilla clarita, sin borde negro, linea de color."""
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.majorGridlines = _light_gridlines()
    chart.x_axis.majorGridlines = None
    if n:
        chart.x_axis.tickLblSkip = max(1, n // 7)
        chart.x_axis.tickMarkSkip = max(1, n // 7)
    # quitar el borde grueso del area de la grafica
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.line = LineProperties(noFill=True)
    chart.legend = None
    if chart.series:
        s = chart.series[0]
        s.graphicalProperties.line.solidFill = series_color
        s.graphicalProperties.line.width = 22000
        s.smooth = True


def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except (ValueError, AttributeError):
        return None


def to_local(dt_utc, tz_offset):
    """Aplica el offset de zona (p.ej. '-06:00') a un datetime UTC -> hora local."""
    if dt_utc is None:
        return None
    if tz_offset and len(tz_offset) >= 3:
        try:
            sign = -1 if tz_offset[0] == "-" else 1
            hh = int(tz_offset[1:3])
            mm = int(tz_offset[4:6]) if len(tz_offset) >= 6 else 0
            return dt_utc + sign * timedelta(hours=hh, minutes=mm)
        except (ValueError, IndexError):
            return dt_utc
    return dt_utc


def fmt_dur(minutes):
    if not minutes:
        return ""
    h = int(minutes // 60)
    m = int(round(minutes % 60))
    return f"{h}h{m:02d}m" if h else f"{m}m"


def load_records(prefix):
    files = sorted(glob.glob(os.path.join(DATA_DIR, f"{prefix}_*.json")))
    if not files:
        return []
    with open(files[-1], encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and isinstance(data.get("records"), list):
        return data["records"]
    if isinstance(data, list):
        return data
    return [data] if isinstance(data, dict) else []


def load_single(prefix):
    files = sorted(glob.glob(os.path.join(DATA_DIR, f"{prefix}_*.json")))
    if not files:
        return {}
    with open(files[-1], encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}


def kcal(kilojoule):
    return round(kilojoule * 0.239006, 0) if isinstance(kilojoule, (int, float)) else None


def _avg(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def acwr_color(a):
    """Zonas Garmin: 0.8-1.3 optimo (verde), 1.3-1.5 o <0.8 (amarillo), >1.5 riesgo (rojo)."""
    if not isinstance(a, (int, float)):
        return SLATE
    if 0.8 <= a <= 1.3:
        return GOOD
    if a > 1.5:
        return BAD
    return WARN


def pmc_series(cyc):
    """Curva de forma: Fitness (CTL, 42d), Fatiga (ATL, 7d), Forma (TSB=CTL_ayer-ATL_ayer).
    Usa el strain diario como carga. Devuelve lista de (fecha, carga, ctl, atl, tsb)."""
    loads = [(x["fecha"], x["strain"]) for x in cyc if isinstance(x.get("strain"), (int, float))]
    if not loads:
        return []
    init = sum(v for _, v in loads[:7]) / min(7, len(loads))  # semilla para no arrancar en 0
    c = a = prev_c = prev_a = init
    out = []
    for d, l in loads:
        c = c + (l - c) / 42
        a = a + (l - a) / 7
        out.append((d, l, round(c, 1), round(a, 1), round(prev_c - prev_a, 1)))
        prev_c, prev_a = c, a
    return out


def _num(v):
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _manual_date(s):
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return parse_dt(s)


def load_manual():
    """Registro manual de natacion: datos/registro-natacion.csv
    Columnas: fecha, km_natacion, sesiones_nado, min_pesas, notas
    (WHOOP no mide distancia en alberca; este es el volumen REAL de nado.)"""
    path = os.path.join(DATA_DIR, "registro-natacion.csv")
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            fecha = _manual_date(r.get("fecha") or r.get("Fecha"))
            if not fecha:
                continue
            rows.append({
                "fecha": fecha,
                "km": _num(r.get("km_natacion")),
                "sesiones": _num(r.get("sesiones_nado")),
                "min_pesas": _num(r.get("min_pesas")),
                "notas": (r.get("notas") or "").strip(),
            })
    return sorted(rows, key=lambda x: x["fecha"])


# ---------- Extraccion ----------
def extract_recovery():
    rows = []
    for r in load_records("recovery"):
        sc = r.get("score") or {}
        rows.append({
            "fecha": parse_dt(r.get("created_at")),
            "recovery": sc.get("recovery_score"),
            "hrv": sc.get("hrv_rmssd_milli"),
            "rhr": sc.get("resting_heart_rate"),
            "spo2": sc.get("spo2_percentage"),
            "temp": sc.get("skin_temp_celsius"),
        })
    return sorted([x for x in rows if x["fecha"]], key=lambda x: x["fecha"])


def extract_sleep():
    rows = []
    for r in load_records("sueno"):
        sc = r.get("score") or {}
        stage = sc.get("stage_summary") or {}
        in_bed = stage.get("total_in_bed_time_milli")
        rows.append({
            "fecha": parse_dt(r.get("start")),
            "rendimiento": sc.get("sleep_performance_percentage"),
            "eficiencia": sc.get("sleep_efficiency_percentage"),
            "consistencia": sc.get("sleep_consistency_percentage"),
            "resp": sc.get("respiratory_rate"),
            "horas": round(in_bed / 3_600_000, 2) if isinstance(in_bed, (int, float)) else None,
        })
    return sorted([x for x in rows if x["fecha"]], key=lambda x: x["fecha"])


def extract_cycles():
    rows = []
    for r in load_records("cycles"):
        sc = r.get("score") or {}
        rows.append({
            "fecha": parse_dt(r.get("start")),
            "strain": sc.get("strain"),
            "fc_prom": sc.get("average_heart_rate"),
            "fc_max": sc.get("max_heart_rate"),
            "kcal": kcal(sc.get("kilojoule")),
        })
    return sorted([x for x in rows if x["fecha"]], key=lambda x: x["fecha"])


def extract_workouts():
    rows = []
    for r in load_records("workouts"):
        sc = r.get("score") or {}
        dist = sc.get("distance_meter")
        tz = r.get("timezone_offset")
        start = to_local(parse_dt(r.get("start")), tz)
        end = to_local(parse_dt(r.get("end")), tz)
        dur = (end - start).total_seconds() / 60 if (start and end) else None
        rows.append({
            "fecha": start,
            "inicio": start,
            "fin": end,
            "dur_min": dur,
            "deporte": r.get("sport_name") or (f"sport {r.get('sport_id')}" if r.get("sport_id") is not None else "—"),
            "strain": sc.get("strain"),
            "fc_prom": sc.get("average_heart_rate"),
            "fc_max": sc.get("max_heart_rate"),
            "kcal": kcal(sc.get("kilojoule")),
            "km": round(dist / 1000, 2) if isinstance(dist, (int, float)) else None,
        })
    return sorted([x for x in rows if x["fecha"]], key=lambda x: x["fecha"], reverse=True)


# ---------- Estilos de tabla ----------
def style_table(ws, headers, rows, start_row=1, num_formats=None):
    """Escribe una tabla con encabezado oscuro y filas alternadas. rows = lista de listas."""
    num_formats = num_formats or {}
    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.fill = _fill(INK)
        c.font = Font(bold=True, color=WHITE, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[start_row].height = 24

    for i, row in enumerate(rows):
        r = start_row + 1 + i
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            if i % 2 == 1:
                c.fill = _fill(ROW_ALT)
            if j in num_formats:
                c.number_format = num_formats[j]
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    for j, h in enumerate(headers, start=1):
        width = 22 if j == 1 else max(12, len(str(h)) + 4)
        ws.column_dimensions[get_column_letter(j)].width = width


def recovery_zone_color(v):
    if not isinstance(v, (int, float)):
        return PANEL
    if v >= 67:
        return GOOD
    if v >= 34:
        return WARN
    return BAD


def traffic_light(ws, col, n, g, r, higher_better=True):
    """Semaforo pastel por celda en la columna `col`, filas 2..n+1.
    higher_better: verde >= g, rojo < r, amarillo entre.  (g > r)
    lower_better:  verde <= g, rojo > r, amarillo entre.  (g < r)"""
    if n < 1 or g is None or r is None:
        return
    rng = f"{col}2:{col}{n + 1}"
    eps = 0.001
    green, yellow, red = _fill(GREEN_SOFT), _fill(YELLOW_SOFT), _fill(RED_SOFT)
    if higher_better:
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=[str(g)], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=[str(r)], fill=red))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(r), str(g - eps)], fill=yellow))
    else:
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThanOrEqual", formula=[str(g)], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=[str(r)], fill=red))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(g + eps), str(r)], fill=yellow))


# ---------- Construccion del dashboard ----------
def build():
    perfil = load_single("perfil")
    rec = extract_recovery()
    slp = extract_sleep()
    cyc = extract_cycles()
    wko = extract_workouts()

    if not any([rec, slp, cyc, wko]):
        sys.exit(
            f"\nNo encontre datos en {DATA_DIR}.\n"
            "Corre primero:  python whoop_sync.py"
        )

    wb = Workbook()

    # ===== Hoja Dashboard =====
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Anchos: A margen; B..K contenido
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 11.5

    # Barra de acento
    for col in range(2, 12):
        ws.cell(row=1, column=col).fill = _fill(ACCENT)
    ws.row_dimensions[1].height = 6

    # Titulo
    ws.merge_cells("B2:K3")
    t = ws["B2"]
    t.value = "TID-MAX  ·  Panel WHOOP"
    t.fill = _fill(INK)
    t.font = Font(bold=True, color=WHITE, size=24)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 12):
        ws.cell(row=2, column=col).fill = _fill(INK)
        ws.cell(row=3, column=col).fill = _fill(INK)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # Subtitulo
    ws.merge_cells("B4:K4")
    nombre = f"{perfil.get('first_name', '')} {perfil.get('last_name', '')}".strip() or "Atleta"
    st = ws["B4"]
    st.value = f"Atleta: {nombre}    ·    Generado: {datetime.now():%d/%b/%Y %H:%M}    ·    Fuente: WHOOP API v2"
    st.font = Font(color="5A6B7B", size=10, italic=True)
    st.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 20

    # Ultimos valores
    last_rec = rec[-1] if rec else {}
    last_slp = slp[-1] if slp else {}
    last_cyc = cyc[-1] if cyc else {}

    man = load_manual()
    last_man = man[-1] if man else {}
    if man:
        ref_date = man[-1]["fecha"]
        km_week = round(sum(m["km"] for m in man if m.get("km") and (ref_date - m["fecha"]).days < 7), 1)
    else:
        km_week = None

    cards = [
        ("RECOVERY", last_rec.get("recovery"), '0"%"', recovery_zone_color(last_rec.get("recovery"))),
        ("HRV (rMSSD)", last_rec.get("hrv"), '0.0" ms"', ACCENT),
        ("FC EN REPOSO", last_rec.get("rhr"), '0" lpm"', BLUE),
        ("SUEÑO", last_slp.get("rendimiento"), '0"%"', INDIGO),
        ("STRAIN HOY", last_cyc.get("strain"), '0.0', ORANGE),
    ]

    # Tarjetas KPI: cada una ocupa 2 columnas (B:C, D:E, F:G, H:I, J:K)
    start_cols = [2, 4, 6, 8, 10]
    label_row, value_row = 6, 7
    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 40
    ws.row_dimensions[8].height = 8
    for (label, value, fmt, color), col in zip(cards, start_cols):
        c1, c2 = get_column_letter(col), get_column_letter(col + 1)
        ws.merge_cells(f"{c1}{label_row}:{c2}{label_row}")
        lc = ws[f"{c1}{label_row}"]
        lc.value = label
        lc.fill = _fill(color)
        lc.font = Font(bold=True, color=WHITE, size=9)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{c1}{value_row}:{c2}{value_row}")
        vc = ws[f"{c1}{value_row}"]
        vc.value = value if value is not None else "—"
        vc.fill = _fill(color)
        vc.font = Font(bold=True, color=WHITE, size=26)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(value, (int, float)):
            vc.number_format = fmt
        # borde inferior de acento visual
        for cc in (c1, c2):
            ws[f"{cc}{label_row}"].border = Border(top=Side(style="thin", color=color))

    # ===== Banda NATACION (volumen real, registro manual) =====
    ws.merge_cells("B9:K9")
    nh = ws["B9"]
    nh.value = "NATACIÓN · volumen real (registro manual)    —    WHOOP no mide distancia en alberca (sin GPS)"
    if not man:
        nh.value += "    ·    crea datos/registro-natacion.csv"
    nh.font = Font(bold=True, color=CYAN, size=10)
    nh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[9].height = 18

    swim_cards = [
        ("KM NADO (ÚLT.)", last_man.get("km"), '0.0" km"', CYAN),
        ("KM 7 DÍAS", km_week, '0.0" km"', AQUA),
        ("SESIONES (ÚLT.)", last_man.get("sesiones"), "0", TEAL),
        ("MIN PESAS (ÚLT.)", last_man.get("min_pesas"), '0" min"', SLATE),
    ]
    ws.row_dimensions[10].height = 16
    ws.row_dimensions[11].height = 34
    ws.row_dimensions[12].height = 10
    for (label, value, fmt, color), col in zip(swim_cards, [2, 4, 6, 8]):
        c1, c2 = get_column_letter(col), get_column_letter(col + 1)
        ws.merge_cells(f"{c1}10:{c2}10")
        lc = ws[f"{c1}10"]
        lc.value = label
        lc.fill = _fill(color)
        lc.font = Font(bold=True, color=WHITE, size=8.5)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{c1}11:{c2}11")
        vc = ws[f"{c1}11"]
        vc.value = value if value is not None else "—"
        vc.fill = _fill(color)
        vc.font = Font(bold=True, color=WHITE, size=20)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(value, (int, float)):
            vc.number_format = fmt

    # ===== Hojas de detalle (datos para tablas y graficas) =====
    ws_rec = wb.create_sheet("Recovery")
    style_table(
        ws_rec,
        ["Fecha", "Recovery %", "HRV (ms)", "FC reposo", "SpO2 %", "Temp piel °C"],
        [[x["fecha"].strftime("%d/%b") if x["fecha"] else "",
          x["recovery"], x["hrv"], x["rhr"], x["spo2"], x["temp"]] for x in rec],
        num_formats={2: '0"%"', 3: "0.0", 4: "0", 5: '0"%"', 6: "0.0"},
    )
    # Semaforo pastel por dia, ajustado a la linea base de Gael
    if rec:
        n_rec = len(rec)
        hrv_vals = [x["hrv"] for x in rec if isinstance(x["hrv"], (int, float))]
        rhr_vals = [x["rhr"] for x in rec if isinstance(x["rhr"], (int, float))]
        hrv_base = sum(hrv_vals) / len(hrv_vals) if hrv_vals else None
        rhr_base = sum(rhr_vals) / len(rhr_vals) if rhr_vals else None
        # Recovery %: zonas WHOOP (verde >=67, amarillo 34-66, rojo <34)
        traffic_light(ws_rec, "B", n_rec, 67, 34, higher_better=True)
        # HRV: mas alta = mejor. Bajones >~20% de su base = bandera roja.
        if hrv_base:
            traffic_light(ws_rec, "C", n_rec, round(hrv_base * 0.90, 1), round(hrv_base * 0.78, 1), higher_better=True)
        # FC en reposo: mas baja = mejor. Elevada sobre su base = riesgo.
        if rhr_base:
            traffic_light(ws_rec, "D", n_rec, round(rhr_base * 1.05, 1), round(rhr_base * 1.12, 1), higher_better=False)

    ws_slp = wb.create_sheet("Sueno")
    style_table(
        ws_slp,
        ["Fecha", "Rendimiento %", "Eficiencia %", "Consistencia %", "Resp (rpm)", "Horas en cama"],
        [[x["fecha"].strftime("%d/%b") if x["fecha"] else "",
          x["rendimiento"], x["eficiencia"], x["consistencia"], x["resp"], x["horas"]] for x in slp],
        num_formats={2: '0"%"', 3: '0"%"', 4: '0"%"', 5: "0.0", 6: "0.0"},
    )
    if slp:
        n_slp = len(slp)
        traffic_light(ws_slp, "B", n_slp, 85, 70, higher_better=True)   # Rendimiento del sueno
        traffic_light(ws_slp, "F", n_slp, 7, 6, higher_better=True)     # Horas en cama

    ws_cyc = wb.create_sheet("Strain")
    style_table(
        ws_cyc,
        ["Fecha", "Strain", "FC prom", "FC max", "Kcal"],
        [[x["fecha"].strftime("%d/%b") if x["fecha"] else "",
          x["strain"], x["fc_prom"], x["fc_max"], x["kcal"]] for x in cyc],
        num_formats={2: "0.0", 3: "0", 4: "0", 5: "0"},
    )

    ws_wko = wb.create_sheet("Workouts")
    style_table(
        ws_wko,
        ["Fecha", "Deporte", "Inicio", "Fin", "Duración", "Strain", "FC prom", "FC max", "Kcal", "Km (WHOOP)*"],
        [[x["fecha"].strftime("%d/%b/%Y") if x["fecha"] else "",
          x["deporte"],
          x["inicio"].strftime("%H:%M") if x["inicio"] else "",
          x["fin"].strftime("%H:%M") if x["fin"] else "",
          fmt_dur(x["dur_min"]),
          x["strain"], x["fc_prom"], x["fc_max"], x["kcal"], x["km"]] for x in wko],
        num_formats={6: "0.0", 7: "0", 8: "0", 9: "0", 10: "0.00"},
    )
    note_row = ws_wko.max_row + 2
    ws_wko.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    nc = ws_wko.cell(row=note_row, column=1)
    nc.value = ("*  En natación el Km de WHOOP NO es confiable (sin GPS en alberca / no cuenta vueltas). "
                "El volumen real está en la hoja 'Natación'.")
    nc.font = Font(italic=True, color=BAD, size=10)
    nc.alignment = Alignment(horizontal="left", vertical="center")

    # ===== Hoja Natacion (volumen real) =====
    ws_nat = wb.create_sheet("Natación")
    if man:
        style_table(
            ws_nat,
            ["Fecha", "Km nado", "Sesiones", "Min pesas", "Notas"],
            [[m["fecha"].strftime("%d/%b/%Y"), m.get("km"), m.get("sesiones"),
              m.get("min_pesas"), m.get("notas")] for m in reversed(man)],
            num_formats={2: '0.0', 3: "0", 4: "0"},
        )
        ws_nat.column_dimensions[get_column_letter(5)].width = 34
        # Grafica de barras: km de nado por dia (cronologico)
        nrows = len(man)
        # datos cronologicos en columnas auxiliares (H=fecha, I=km) para la grafica
        ws_nat.cell(row=1, column=8, value="Fecha")
        ws_nat.cell(row=1, column=9, value="Km")
        for i, m in enumerate(man, start=2):
            ws_nat.cell(row=i, column=8, value=m["fecha"].strftime("%d/%b"))
            ws_nat.cell(row=i, column=9, value=m.get("km"))
        ws_nat.column_dimensions[get_column_letter(8)].hidden = True
        ws_nat.column_dimensions[get_column_letter(9)].hidden = True
        bc = BarChart()
        bc.type = "col"
        bc.title = "Km de nado por día (real)"
        bc.height = 8
        bc.width = 22
        bc.legend = None
        d = Reference(ws_nat, min_col=9, min_row=1, max_row=nrows + 1)
        cx = Reference(ws_nat, min_col=8, min_row=2, max_row=nrows + 1)
        bc.add_data(d, titles_from_data=True)
        bc.set_categories(cx)
        bc.y_axis.delete = False
        bc.x_axis.delete = False
        bc.y_axis.majorGridlines = _light_gridlines()
        if bc.series:
            bc.series[0].graphicalProperties.solidFill = CYAN
        ws_nat.add_chart(bc, "G2")
    else:
        ws_nat["A1"] = "Sin registro de natación todavía."
        ws_nat["A1"].font = Font(bold=True, size=13, color=INK)
        ws_nat["A3"] = ("Cómo activarlo:  copia el archivo  registro-natacion.ejemplo.csv  a  "
                        "datos/registro-natacion.csv  y anota por día: km_natacion, sesiones_nado, min_pesas.")
        ws_nat["A3"].font = Font(size=11, color="5A6B7B")
        ws_nat.column_dimensions["A"].width = 110

    # ===== Graficas de tendencia en el Dashboard =====
    n = len(rec)
    if n >= 2:
        first_data = max(2, n - 29)  # ultimos ~30 puntos
        cats = Reference(ws_rec, min_col=1, min_row=first_data, max_row=n + 1)

        npts = n + 1 - first_data + 1  # puntos mostrados en el eje X

        chart_rec = LineChart()
        chart_rec.title = "Recovery % (tendencia)"
        chart_rec.height = 7.5
        chart_rec.width = 15
        chart_rec.style = 2
        data = Reference(ws_rec, min_col=2, min_row=1, max_row=n + 1)
        chart_rec.add_data(data, titles_from_data=True)
        chart_rec.set_categories(cats)
        chart_rec.y_axis.scaling.min = 0
        chart_rec.y_axis.scaling.max = 100
        style_chart(chart_rec, npts, GOOD)
        ws.add_chart(chart_rec, "B13")

        chart_hrv = LineChart()
        chart_hrv.title = "HRV rMSSD (ms)"
        chart_hrv.height = 7.5
        chart_hrv.width = 15
        chart_hrv.style = 2
        data_h = Reference(ws_rec, min_col=3, min_row=1, max_row=n + 1)
        chart_hrv.add_data(data_h, titles_from_data=True)
        chart_hrv.set_categories(cats)
        style_chart(chart_hrv, npts, ACCENT)
        ws.add_chart(chart_hrv, "G13")

    # ===== Leyenda del semaforo (aplica a las hojas de detalle) =====
    ws["B30"] = "Semáforo:"
    ws["B30"].font = Font(bold=True, size=9, color=INK)
    ws["B30"].alignment = Alignment(horizontal="left", vertical="center")
    for col, txt, color in [("C", "bien", GREEN_SOFT), ("D", "vigilar", YELLOW_SOFT), ("E", "riesgo", RED_SOFT)]:
        c = ws[f"{col}30"]
        c.value = txt
        c.fill = _fill(color)
        c.font = Font(size=9, color="333333")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("F30:K30")
    ws["F30"] = "en las hojas Recovery y Sueño, ajustado a la línea base de Gael (Recovery, HRV, FC reposo, sueño)"
    ws["F30"].font = Font(size=9, italic=True, color="5A6B7B")
    ws["F30"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[30].height = 16

    wb.save(OUT_PATH)
    print("\n===================== LISTO =====================")
    print(f"Dashboard generado: {OUT_PATH}")
    print(f"Recovery: {len(rec)} dias  |  Sueno: {len(slp)}  |  Strain: {len(cyc)}  |  Workouts: {len(wko)}  |  Natacion (manual): {len(man)}")
    print("\nAbrelo con:")
    print(f'    open "{OUT_PATH}"')
    print("=================================================\n")


if __name__ == "__main__":
    build()
